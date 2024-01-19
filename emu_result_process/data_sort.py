import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def style_cells(cell, o_column_value):
    # 根据 'full_RMSU_err' 列的值设置颜色
    if -5 <= o_column_value <= 5:
        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # 绿色
    else:
        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # 红色

def filter_and_style_to_sheets(excel_file_path, output_file_path, dynamic_column_name, filter_column_name, filter_condition):
    # 读取Excel文件的第一行作为列名
    df = pd.read_excel(excel_file_path, header=0)

    # 获取动态列（dynamic_test）的所有唯一值
    unique_values = df[dynamic_column_name].unique()

    # 使用 Openpyxl 创建一个 Workbook 对象
    workbook = Workbook()

    # 遍历每个动态列的唯一值进行筛选并保存结果到不同的 sheet
    for value_to_filter in unique_values:
        # 先进行动态列的筛选
        dynamic_filtered_data = df[df[dynamic_column_name] == value_to_filter]

        # 然后在动态列的基础上进行静态列（full_RMSU_err）的筛选
        static_filtered_data = dynamic_filtered_data[dynamic_filtered_data[filter_column_name].apply(filter_condition)]

        # 创建一个 sheet，sheet 名称为动态列的唯一值
        sheet = workbook.create_sheet(title=str(value_to_filter))

        # 将数据写入 sheet
        for r_idx, row in enumerate(pd.DataFrame(static_filtered_data).values):
            for c_idx, value in enumerate(row):
                cell = sheet.cell(row=r_idx + 2, column=c_idx + 1, value=value)
                if filter_column_name in df.columns:
                    o_column_value = static_filtered_data.iloc[r_idx][filter_column_name]
                    style_cells(cell, o_column_value)

    # 删除默认创建的 sheet
    del workbook["Sheet"]

    # 保存 Workbook 到文件
    workbook.save(output_file_path)
    print(f"所有筛选结果以及单元格样式保存到 {output_file_path} 的不同 sheet 中。")

# 定义筛选条件，例如，选择 'full_RMSU_err' 列中大于等于 5 的行
filter_condition = lambda x: x >= 5

# 使用函数进行调用
excel_file_path = r"C:\Users\huadi\Desktop\gitcode\library\emu_result_process\data\result_ori.xlsx"
output_file_path = r"C:\Users\huadi\Desktop\gitcode\library\emu_result_process\data\out_styled_and_filtered.xlsx"
dynamic_column_name = 'dynamic_test'
filter_column_name = 'full_RMSU_err'

filter_and_style_to_sheets(excel_file_path, output_file_path, dynamic_column_name, filter_column_name, filter_condition)
